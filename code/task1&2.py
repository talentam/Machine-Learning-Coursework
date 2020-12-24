from openpyxl import load_workbook, Workbook
from collections import Counter
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import pearsonr
from minepy import MINE
import scipy.stats as stats

from scipy.spatial.distance import correlation

###################################
# instruction to run the code
# the root path to run is task1&2.py
# in other words, "./task1&2.py" can access task1&2.py
###################################


# read worksheet and find out the overlapped year
def read_worksheet(sheetName):
    # store the all the data in the list
    data_list = []
    # this list used for finding the depth that appears most
    depth_list = []
    start_year = 3000
    end_year = 1000
    for row in range(2, sheetName.max_row + 1):
        column4 = sheetName.cell(row=row, column=4).value
        column5 = sheetName.cell(row=row, column=5).value
        column6 = sheetName.cell(row=row, column=6).value
        column7 = sheetName.cell(row=row, column=7).value
        temporary_list = [column4, column5, column6, column7]
        data_list.append(temporary_list)
        depth_list.append(column6)

        if column5.year < start_year:
            start_year = column5.year
        elif column5.year > end_year:
            end_year = column5.year

    return data_list, depth_list, start_year, end_year


# read workbook
def read_workbook(path):
    workbook = load_workbook(path)
    Chla_ws = workbook['CHLA ']
    Temp_ws = workbook['TEMPERATURE']
    TotalP_ws = workbook['Total P ']
    Chla_data, Chla_depth, start_year1, end_year1 = read_worksheet(Chla_ws)
    Temp_data, Temp_depth, start_year2, end_year2 = read_worksheet(Temp_ws)
    TotalP_data, TotalP_depth, start_year3, end_year3 = read_worksheet(TotalP_ws)

    # find out the start year ,end year
    start_year_max = max(start_year1, start_year2, start_year3)
    end_year_min = min(end_year1, end_year2, end_year3)

    # find out the majority depth
    depth_total = Chla_depth + Temp_depth + TotalP_depth
    max_depth = Counter(depth_total).most_common(1)[0][0]

    return Chla_data, Temp_data, TotalP_data, start_year_max, end_year_min, max_depth


# remove the rows that some cell is empty or the year is not overlapped or month exceed May-October
def data_cleaning(data):
    remove_index = []
    for i, value in enumerate(data):
        length = len(value)
        if value[length-1] is None or value[length-2] is None or value[length-3].year < starting_year \
                or value[length-3].year > ending_year or value[length-4] == 2 \
                or value[length-3].month < starting_month or value[length-3].month > ending_month \
                or value[length-2] != depth:
            remove_index.append(i)

    remove_index.reverse()
    for index in remove_index:
        data.pop(index)


# create empty list to store the data
def initializeEmptyList():
    rows = ending_year - starting_year + 1
    columns = ending_month - starting_month + 1
    empty_list = [[[] for i in range(columns)] for j in range(rows)]
    return empty_list


# computer the data which does not need to be complete (calculate mean value for existing data)
def averageData(original_list):
    empty_list = initializeEmptyList()
    year_num = ending_year - starting_year + 1
    month_num = ending_month - starting_month + 1

    for row in original_list:
        year = row[1].year - starting_year
        month = row[1].month - starting_month
        empty_list[year][month].append(row[3])

    for i in range(year_num):
        for j in range(month_num):
            if len(empty_list[i][j]) == 0:
                empty_list[i][j] = [0]
            else:
                empty_list[i][j] = [np.mean(empty_list[i][j])]
    return empty_list


# count zero numbers in a list
def countZero(input_list):
    zero_num = 0
    for i in input_list:
        if i[0] == 0:
            zero_num = zero_num + 1
    return zero_num


def bestMatch(list1, list2, list3, orig_list1, orig_list2, orig_list3):
    previous_year = 0
    previous_month = 0
    previous_day = 0
    empty_list_chla = initializeEmptyList()
    empty_list_temp = initializeEmptyList()
    empty_list_totalP = initializeEmptyList()
    for row in orig_list1:
        year = row[1].year
        month = row[1].month
        day = row[1].day
        # prevent reading the same day twice for temp and totalP
        if year == previous_year and month == previous_month and day == previous_day:
            continue
        else:
            previous_year = year
            previous_month = month
            previous_day = day

            index1 = []
            index2 = []
            for i, data in enumerate(orig_list2):
                # match for temp
                if data[1].year == year and data[1].month == month and data[1].day == day:
                    index1.append(i)
                if data[1].year > year:
                    break

            for i, data in enumerate(orig_list3):
                # match for totalP
                if data[1].year == year and data[1].month == month and data[1].day == day:
                    index2.append(i)
                if data[1].year > year:
                    break
            # find best match
            if len(index1) != 0 and len(index2) != 0:
                empty_list_chla[year - starting_year][month - starting_month].append(row[3])
                for index in index1:
                    empty_list_temp[orig_list2[index][1].year - starting_year][orig_list2[index][1].month - starting_month].append(orig_list2[index][3])
                for index in index2:
                    empty_list_totalP[orig_list3[index][1].year - starting_year][orig_list3[index][1].month - starting_month].append(orig_list3[index][3])

    # calculate the mean of best match
    year_num = ending_year - starting_year + 1
    month_num = ending_month - starting_month + 1
    for i in range(year_num):
        for j in range(month_num):
            if len(empty_list_chla[i][j]) == 0:
                empty_list_chla[i][j] = [0]
            else:
                empty_list_chla[i][j] = [np.mean(empty_list_chla[i][j])]
            if len(empty_list_temp[i][j]) == 0:
                empty_list_temp[i][j] = [0]
            else:
                empty_list_temp[i][j] = [np.mean(empty_list_temp[i][j])]
            if len(empty_list_totalP[i][j]) == 0:
                empty_list_totalP[i][j] = [0]
            else:
                empty_list_totalP[i][j] = [np.mean(empty_list_totalP[i][j])]

    # update best match
    for i in range(year_num):
        for j in range(month_num):
            if empty_list_chla[i][j][0] != 0 and empty_list_chla[i][j][0] != list1[i][j][0]:
                list1[i][j][0] = empty_list_chla[i][j][0]
            if empty_list_temp[i][j][0] != 0 and empty_list_temp[i][j][0] != list2[i][j][0]:
                list2[i][j][0] = empty_list_temp[i][j][0]
            if empty_list_totalP[i][j][0] != 0 and empty_list_totalP[i][j][0] != list3[i][j][0]:
                list3[i][j][0] = empty_list_totalP[i][j][0]


# method 1: mean value to complete missing data
def meanCalculation(input_list):
    month_num = ending_month - starting_month + 1
    for year in input_list:
        # count the zero number
        zero_num = countZero(year)
        while 0 < zero_num < 4:
            # check condition of 101
            for i in range(0, month_num-2):
                if year[i][0] != 0 and year[i+1][0] == 0 and year[i+2][0] != 0:
                    year[i + 1][0] = max((year[i][0] + year[i + 2][0])/2, 0)
                    zero_num = zero_num - 1
                    if zero_num == 0:
                        break

            # check condition of 110 or 011
            for i in range(0, month_num - 2):
                # 110
                if year[i][0] != 0 and year[i + 1][0] != 0 and year[i + 2][0] == 0:
                    year[i + 2][0] = max(2*year[i + 1][0] - year[i][0], 0)
                    zero_num = zero_num - 1
                    break
                # 011
                elif year[i][0] == 0 and year[i + 1][0] != 0 and year[i + 2][0] != 0:
                    year[i][0] = max(2 * year[i + 1][0] - year[i+2][0], 0)
                    zero_num = zero_num - 1
                    break
    return input_list


# method 2: polynomial regression to complete missing data
def polynomial(input_list, degree, max_value):
    plt.figure(figsize=(30, 24))
    for i, year in enumerate(input_list):
        # skip the year which did not have data
        if countZero(year) == 6:
            continue
        x = []
        y = []
        for j, month in enumerate(year):
            if month[0] != 0:
                x.append(j + starting_month)
                y.append(month[0])

        # polynomial regression
        a = np.polyfit(x, y, degree)
        fx = np.poly1d(a)

        # use the regression function to complete the missing data
        for j, month in enumerate(year):
            if month[0] == 0:
                input_list[i][j] = [max(fx(j + starting_month), 0)]

        # plot the regression function

        plt.subplot(4, 4, i + 1)
        plt.scatter(x, y, color='black')
        plt.xticks(fontsize=20)
        plt.yticks(fontsize=20)
        plt.plot(np.linspace(starting_month, ending_month, 100), fx(np.linspace(starting_month, ending_month, 100)), 'r-', lw=3)
        plt.xlim(starting_month-0.3, ending_month+0.3)
        plt.ylim(0, max_value)

    plt.show()
    return input_list


# output the complete excel file
def outputTable(list1, list2, list3, info_list, table_name):
    sheet = wb.create_sheet(table_name)
    sheet.append(['MIDAS', 'LAKE', 'Town(s)', 'STATION', 'Date', 'DEPTH', 'CHLA (mg/L)', 'TEMPERATURE (Centrigrade)', 'Total P (mg/L)'])
    for i in range(0, len(list1) * len(list1[0])):
        if countZero(list1[i // 6]) < 6:
            # append normal information
            row = ['5448', 'China Lake', 'China, Vassalboro']
            # append station
            row.append(info_list[0][0])
            # append year/month
            year = i // 6 + starting_year
            month = i % 6 + starting_month
            year_month = str(year)+'/'+str(month)
            row.append(year_month)
            # append depth
            row.append(info_list[0][2])
            # append CHLA
            row.append(round(list1[i // 6][i % 6][0], 6))
            # append TEMPERATURE
            row.append(round(list2[i // 6][i % 6][0], 6))
            # append TotalP
            row.append(round(list3[i // 6][i % 6][0], 6))

            sheet.append(row)


# change the format of the data for the correlation calculation
def data_preprocessing(list1, list2, list3):
    x = []
    y = []
    z = []
    skip_year = []
    for i, year in enumerate(list1):
        if countZero(year) == 6:
            skip_year.append(i)
            continue
        else:
            for j, month in enumerate(year):
                x.append(round(month[0], 4))

    for i, year in enumerate(list2):
        if i not in skip_year:
            for j, month in enumerate(year):
                y.append(round(month[0], 1))

    for i, year in enumerate(list3):
        if i not in skip_year:
            for j, month in enumerate(year):
                z.append(round(month[0], 3))

    return x, y, z


def covariance(variable1, variable2):
    covXY = np.cov(variable1, variable2)[1][0]
    return covXY


def pearson(variable1, variable2):
    pearson_coefficient = pearsonr(variable1, variable2)[0]
    return pearson_coefficient


def mine(variable1, variable2):
    m = MINE()
    m.compute_score(variable1, variable2)
    return m.mic()


def spearman(variable1, variable2):
    return stats.spearmanr(variable1, variable2)[0]


def distcorr(variable1, variable2):
    return correlation(variable1, variable2)


def printRanking(method, temp, totalp):
    print(method + ': temperature (' + str(temp) + '); TotalP (' + str(totalp) + ')')
    if temp > totalp:
        print('     temperature is more important ')
    else:
        print('     TotalP is more important ')



# read the workbook and find overlapped years
print('[INFO] reading workbook')
Chla_list, Temp_list, TotalP_list, starting_year, ending_year, depth = read_workbook('./lake_data/China lake.xlsx')

# the rows contains empty cell will be removed
# Also, the year outside the overlapped period will be removed
# In addition, only 5-10 months will be kept
print('[INFO] cleaning data')
starting_month, ending_month = 5, 10
data_cleaning(Chla_list)
data_cleaning(Temp_list)
data_cleaning(TotalP_list)

# average the data so that one month contains one record
print('[INFO] matching data')
Chla_avg_m1 = averageData(Chla_list)
Temp_avg_m1 = averageData(Temp_list)
TotalP_avg_m1 = averageData(TotalP_list)

Chla_avg_m2 = averageData(Chla_list)
Temp_avg_m2 = averageData(Temp_list)
TotalP_avg_m2 = averageData(TotalP_list)

bestMatch(Chla_avg_m1, Temp_avg_m1, TotalP_avg_m1, Chla_list, Temp_list, TotalP_list)
bestMatch(Chla_avg_m2, Temp_avg_m2, TotalP_avg_m2, Chla_list, Temp_list, TotalP_list)

# method 1: mean value calculation
print('[INFO] mean value calculation')
Chla_mean = meanCalculation(Chla_avg_m1)
Temp_mean = meanCalculation(Temp_avg_m1)
TotalP_mean = meanCalculation(TotalP_avg_m1)

# method 2: polynomial regression
Chla_poly = polynomial(Chla_avg_m2, 1, 0.04)
Temp_poly = polynomial(Temp_avg_m2, 2, 25)
TotalP_poly = polynomial(TotalP_avg_m2, 2, 0.04)

# output table to excel
wb = Workbook()
outputTable(Chla_mean, Temp_mean, TotalP_mean, Chla_list, 'mean value')
outputTable(Chla_poly, Temp_poly, TotalP_poly, Chla_list, 'polynomial regression')
wb.save("./lake_data/completeChinaLake.xlsx")

# task 2: using the meaning completed data
Chla_mean_data, temp_mean_data, TotalP_mean_data = data_preprocessing(Chla_mean, Temp_mean, TotalP_mean)

# Covariance
cov_Chla_temp = covariance(Chla_mean_data, temp_mean_data)
cov_Chla_TotalP = covariance(Chla_mean_data, TotalP_mean_data)
printRanking('Covariance', cov_Chla_temp, cov_Chla_TotalP)

# Pearson correlation coefficient
Pearson_Chla_temp = pearson(Chla_mean_data, temp_mean_data)
Pearson_Chla_TotalP = pearson(Chla_mean_data, TotalP_mean_data)
printRanking('Pearson', Pearson_Chla_temp, Pearson_Chla_TotalP)

# Spearman Correlation Coefficient
spearman_Chla_temp = spearman(Chla_mean_data, temp_mean_data)
spearman_Chla_TotalP = spearman(Chla_mean_data, TotalP_mean_data)
printRanking('Spearman', spearman_Chla_temp, spearman_Chla_TotalP)

# Distance correlation
distcorr_Chla_temp = distcorr(Chla_mean_data, temp_mean_data)
distcorr_Chla_TotalP = distcorr(Chla_mean_data, TotalP_mean_data)
printRanking('Distance correlation', distcorr_Chla_temp, distcorr_Chla_TotalP)

# Maximal information coefficient (MIC)
mine_Chla_temp = mine(Chla_mean_data, temp_mean_data)
mine_Chla_TotalP = mine(Chla_mean_data, TotalP_mean_data)
printRanking('MIC', mine_Chla_temp, mine_Chla_TotalP)
